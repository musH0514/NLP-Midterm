import openpyxl
import os
import csv
from collections import defaultdict
import numpy as np
import matplotlib.pyplot as plt

print("开始更新高活跃度作者定义并重新分析...")

# 1. 从04表中提取作者、年份和论文信息
print("读取04表，提取作者和年份信息...")
excel_file = r'f:\projects\nlp\NLP-Midterm\r1\csv\04_Diaspora_Articles_12195(1).xlsx'

# 使用read_only模式打开Excel文件
workbook = openpyxl.load_workbook(excel_file, read_only=True)
sheet = workbook.active

# 找到需要的列
author_col = None
year_col = None
row = sheet[1]  # 获取表头行
for cell in row:
    if cell.value == 'Diaspora_Author_Names':
        author_col = cell.column - 1  # 转换为0-based索引
    elif cell.value == 'Publication Year':
        year_col = cell.column - 1  # 转换为0-based索引

if author_col is None or year_col is None:
    print("错误：未找到所需的列")
    exit(1)

print(f"Diaspora_Author_Names列在索引 {author_col}")
print(f"Publication Year列在索引 {year_col}")

# 收集每个作者的发表年份
author_years = defaultdict(list)
processed_count = 0
batch_size = 1000
max_rows = sheet.max_row

for row in sheet.iter_rows(min_row=2, max_row=max_rows):
    processed_count += 1
    if processed_count % batch_size == 0:
        print(f"已处理 {processed_count} / {max_rows} 行")
    
    # 获取作者和年份信息
    author_cell = row[author_col].value
    year_cell = row[year_col].value
    
    if author_cell and year_cell:
        # 处理多个作者的情况
        authors = [author.strip() for author in str(author_cell).split(';') if author.strip()]
        year = int(year_cell)
        
        for author in authors:
            author_years[author].append(year)

workbook.close()

print(f"已处理 {processed_count} 行")
print(f"已提取 {len(author_years)} 位作者的发表年份")

# 2. 根据新定义识别高活跃度作者（任何5年区间发表≥5篇）
print("\n根据新定义识别高活跃度作者...")

# 定义检查函数：是否有任何5年区间发表≥5篇
def is_highly_productive(years):
    if len(years) < 5:
        return False
    
    # 对年份进行排序
    sorted_years = sorted(years)
    
    # 检查所有可能的5年区间
    for i in range(len(sorted_years) - 4):
        # 获取当前区间的起始年份
        start_year = sorted_years[i]
        end_year = start_year + 4  # 5年区间（包含起始年份）
        
        # 统计该区间内的论文数
        count = 0
        for year in sorted_years[i:]:
            if year <= end_year:
                count += 1
            else:
                break
        
        if count >= 5:
            return True
    
    return False

# 识别高活跃度作者
highly_productive_authors = []
for author, years in author_years.items():
    if is_highly_productive(years):
        # 计算该作者的总论文数
        total_papers = len(years)
        # 计算发表年份范围
        min_year = min(years)
        max_year = max(years)
        year_range = f"{min_year}-{max_year}"
        # 计算平均每年发表论文数
        if min_year == max_year:
            average = total_papers
        else:
            years_span = max_year - min_year + 1
            average = round(total_papers / years_span, 2)
        
        highly_productive_authors.append({
            'Author': author,
            'Articles': total_papers,
            'Year Range': year_range,
            'Average': average
        })

# 按发表论文数降序排序
highly_productive_authors.sort(key=lambda x: x['Articles'], reverse=True)

print(f"根据新定义，高活跃度作者数量: {len(highly_productive_authors)}")

# 3. 保存新的高活跃度作者表
print("\n保存新的高活跃度作者表...")
high_file = r'f:\projects\nlp\NLP-Midterm\r1\csv\highly_productive_scholars.csv'

with open(high_file, 'w', newline='', encoding='utf-8-sig') as f:
    writer = csv.writer(f)
    # 写入表头
    writer.writerow(['Author', 'Articles', 'Year Range', 'Average'])
    
    # 写入数据
    for author_info in highly_productive_authors:
        writer.writerow([
            author_info['Author'],
            author_info['Articles'],
            author_info['Year Range'],
            author_info['Average']
        ])

print(f"新的高活跃度作者表已保存到: {high_file}")

# 4. 读取作者学科分类表，获取低活跃度作者
print("\n读取作者学科分类表...")
subject_file = r'f:\projects\nlp\NLP-Midterm\r1\csv\unique_authors_with_subject.csv'

author_total_papers = defaultdict(int)
all_authors = set()

with open(subject_file, 'r', encoding='utf-8-sig') as f:
    reader = csv.DictReader(f)
    for row in reader:
        author = row['Author']
        total_papers = int(row['Total_Papers'])
        author_total_papers[author] = total_papers
        all_authors.add(author)

# 创建高活跃度作者集合
high_authors_set = set([author_info['Author'] for author_info in highly_productive_authors])

# 确定低活跃度作者（不在高活跃度作者集合中）
low_authors = [author for author in all_authors if author not in high_authors_set]

print(f"低活跃度作者数量: {len(low_authors)}")
print(f"总作者数: {len(all_authors)}")

# 5. 计算合作论文数（基于完整的04表数据）
print("\n重新计算合作论文数...")

# 重新打开Excel文件计算合作论文数
author_collab_papers = defaultdict(int)

workbook = openpyxl.load_workbook(excel_file, read_only=True)
sheet = workbook.active

processed_count = 0
for row in sheet.iter_rows(min_row=2, max_row=max_rows):
    processed_count += 1
    if processed_count % batch_size == 0:
        print(f"已处理 {processed_count} / {max_rows} 行")
    
    author_cell = row[author_col].value
    if author_cell:
        authors = [author.strip() for author in str(author_cell).split(';') if author.strip()]
        num_authors = len(authors)
        
        if num_authors > 1:  # 合作论文
            for author in authors:
                if author in all_authors:  # 只统计我们关注的作者
                    author_collab_papers[author] += 1

workbook.close()

print(f"已计算 {len(author_collab_papers)} 位作者的合作论文数")

# 6. 计算合作率
print("\n计算合作率...")

# 高活跃度作者合作率
high_collab_rates = []
for author_info in highly_productive_authors:
    author = author_info['Author']
    total = author_info['Articles']
    collab = author_collab_papers.get(author, 0)
    if total > 0:
        high_collab_rates.append(collab / total)

# 低活跃度作者合作率
low_collab_rates = []
for author in low_authors:
    total = author_total_papers.get(author, 0)
    collab = author_collab_papers.get(author, 0)
    if total > 0:
        low_collab_rates.append(collab / total)

print(f"高活跃度作者有效样本: {len(high_collab_rates)} / {len(highly_productive_authors)}")
print(f"低活跃度作者有效样本: {len(low_collab_rates)} / {len(low_authors)}")

# 7. 计算统计指标
print("\n计算统计指标...")

# 高活跃度作者统计
high_stats = {
    'count': len(high_collab_rates),
    'avg_rate': np.mean(high_collab_rates) if high_collab_rates else 0,
    'median_rate': np.median(high_collab_rates) if high_collab_rates else 0
}

# 低活跃度作者统计
low_stats = {
    'count': len(low_collab_rates),
    'avg_rate': np.mean(low_collab_rates) if low_collab_rates else 0,
    'median_rate': np.median(low_collab_rates) if low_collab_rates else 0
}

# 打印统计结果
print("\n=== 高活跃度作者统计 (新定义) ===")
print(f"作者数量: {high_stats['count']}")
print(f"平均合作率: {high_stats['avg_rate']:.4f}")
print(f"中位数合作率: {high_stats['median_rate']:.4f}")

print("\n=== 低活跃度作者统计 ===")
print(f"作者数量: {low_stats['count']}")
print(f"平均合作率: {low_stats['avg_rate']:.4f}")
print(f"中位数合作率: {low_stats['median_rate']:.4f}")

# 8. 生成对比图表
print("\n生成对比图表...")

# 创建图表目录
output_dir = r'f:\projects\nlp\NLP-Midterm\r1\maps\background'
if not os.path.exists(output_dir):
    os.makedirs(output_dir)

# 图表: 平均合作率对比
plt.figure(figsize=(8, 6))

categories = ['Highly Productive', 'Low Productive']
avg_rates = [high_stats['avg_rate'], low_stats['avg_rate']]

bars = plt.bar(categories, avg_rates, color=['#4ECDC4', '#FF6B6B'], width=0.5)

plt.title('Average Collaboration Rate Comparison', fontsize=16)
plt.ylabel('Collaboration Rate', fontsize=12)
plt.ylim(0, 1)

# 在柱子上显示数值
for bar in bars:
    height = bar.get_height()
    plt.text(bar.get_x() + bar.get_width()/2., height + 0.01,
             f'{height:.4f}', ha='center', va='bottom')

# 添加网格线
plt.grid(axis='y', alpha=0.3)

# 保存图表
chart_file = os.path.join(output_dir, 'collaboration_rate_comparison.png')
plt.savefig(chart_file, dpi=300, bbox_inches='tight')
print(f"图表已保存到: {chart_file}")

# 图表: 合作率分布箱线图
plt.figure(figsize=(10, 6))

# 将数据整理为箱线图需要的格式
data_to_plot = [high_collab_rates, low_collab_rates]

# 创建箱线图
box = plt.boxplot(data_to_plot, patch_artist=True, tick_labels=categories)

# 设置箱子颜色
colors = ['#4ECDC4', '#FF6B6B']
for patch, color in zip(box['boxes'], colors):
    patch.set_facecolor(color)
    patch.set_alpha(0.7)

plt.title('Collaboration Rate Distribution', fontsize=16)
plt.ylabel('Collaboration Rate', fontsize=12)
plt.ylim(0, 1)

# 添加网格线
plt.grid(axis='y', alpha=0.3)

# 保存图表
boxplot_file = os.path.join(output_dir, 'collaboration_rate_distribution.png')
plt.savefig(boxplot_file, dpi=300, bbox_inches='tight')
print(f"箱线图已保存到: {boxplot_file}")

# 9. 生成结果表格
print("\n生成结果表格...")

result_file = r'f:\projects\nlp\NLP-Midterm\r1\csv\collaboration_comparison_result.csv'

with open(result_file, 'w', newline='', encoding='utf-8-sig') as f:
    f.write('Metric,Highly Productive Authors,Low Productive Authors\n')
    f.write(f'Number of Authors,{high_stats["count"]},{low_stats["count"]}\n')
    f.write(f'Average Collaboration Rate,{high_stats["avg_rate"]:.4f},{low_stats["avg_rate"]:.4f}\n')
    f.write(f'Median Collaboration Rate,{high_stats["median_rate"]:.4f},{low_stats["median_rate"]:.4f}\n')

print(f"结果表格已保存到: {result_file}")

print("\n=== 更新完成！ ===")
print("\n生成的文件:")
print(f"1. {high_file} - 新定义的高活跃度作者表")
print(f"2. {chart_file} - 平均合作率对比图")
print(f"3. {boxplot_file} - 合作率分布箱线图")
print(f"4. {result_file} - 对比结果表格")
print("\n新定义：高活跃度作者是指在任何5年区间内发表文章≥5篇的作者")
