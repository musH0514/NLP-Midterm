import openpyxl
import csv
import os
from collections import Counter
import re

print("提取关键词对比结果...")

# 定义停用词（与词云生成脚本相同）
stopwords = {
    # 英文常见停用词
    'the', 'and', 'of', 'to', 'a', 'in', 'for', 'is', 'on', 'with', 'as', 'by', 'at', 'from',
    'that', 'this', 'are', 'was', 'were', 'be', 'been', 'being', 'which', 'what', 'who',
    'whom', 'whose', 'when', 'where', 'why', 'how', 'about', 'after', 'before', 'during',
    'between', 'into', 'through', 'over', 'above', 'below', 'up', 'down', 'out', 'off',
    'again', 'further', 'then', 'once', 'here', 'there', 'all', 'any', 'both', 'each',
    'few', 'more', 'most', 'other', 'some', 'such', 'no', 'nor', 'not', 'only', 'own',
    'same', 'so', 'than', 'too', 'very', 'can', 'will', 'just', 'should', 'now', 'but',
    'or', 'because', 'if', 'though', 'while', 'until', 'unless', 'also', 'even', 'however',
    'yet', 'still', 'already', 'almost', 'always', 'never', 'often', 'usually', 'sometimes',
    'may', 'might', 'must', 'shall', 'should', 'would', 'could', 'one', 'two', 'three',
    'four', 'five', 'six', 'seven', 'eight', 'nine', 'ten', 'first', 'second', 'third',
    'fourth', 'fifth', 'many', 'much', 'few', 'little', 'several', 'some', 'any', 'all',
    'none', 'every', 'each', 'either', 'neither', 'such', 'another', 'other', 'others',
    'its', 'it', 'you', 'your', 'yours', 'we', 'our', 'ours', 'they', 'their', 'theirs',
    'he', 'his', 'she', 'her', 'hers', 'me', 'my', 'mine', 'them', 'us', 'i', 'am', 'has',
    'have', 'had', 'do', 'does', 'did', 'doing', 'done', 'go', 'goes', 'went', 'going',
    'gone', 'get', 'gets', 'got', 'getting', 'gotten', 'make', 'makes', 'made', 'making',
    'take', 'takes', 'took', 'taking', 'taken', 'give', 'gives', 'gave', 'giving', 'given',
    'use', 'uses', 'used', 'using', 'see', 'sees', 'saw', 'seeing', 'seen', 'find', 'finds',
    'found', 'finding', 'found', 'look', 'looks', 'looked', 'looking', 'come', 'comes',
    'came', 'coming', 'came', 'bring', 'brings', 'brought', 'bringing', 'brought', 'put',
    'puts', 'put', 'putting', 'let', 'lets', 'let', 'letting', 'keep', 'keeps', 'kept',
    'keeping', 'kept', 'hold', 'holds', 'held', 'holding', 'held', 'say', 'says', 'said',
    'saying', 'said', 'speak', 'speaks', 'spoke', 'speaking', 'spoken', 'tell', 'tells',
    'told', 'telling', 'told', 'know', 'knows', 'knew', 'knowing', 'known', 'think',
    'thinks', 'thought', 'thinking', 'thought', 'believe', 'believes', 'believed',
    'believing', 'believed', 'want', 'wants', 'wanted', 'wanting', 'wanted', 'need',
    'needs', 'needed', 'needing', 'needed', 'like', 'likes', 'liked', 'liking', 'liked',
    'love', 'loves', 'loved', 'loving', 'loved', 'hate', 'hates', 'hated', 'hating',
    'hated', 'work', 'works', 'worked', 'working', 'worked', 'study', 'studies',
    'studied', 'studying', 'studied', 'research', 'researches', 'researched',
    'researching', 'researched', 'analyze', 'analyzes', 'analyzed', 'analyzing',
    'analyzed', 'examine', 'examines', 'examined', 'examining', 'examined', 'test',
    'tests', 'tested', 'testing', 'tested', 'try', 'tries', 'tried', 'trying', 'tried',
    'experiment', 'experiments', 'experimented', 'experimenting', 'experimented',
    'show', 'shows', 'showed', 'showing', 'shown', 'display', 'displays', 'displayed',
    'displaying', 'displayed', 'indicate', 'indicates', 'indicated', 'indicating',
    'indicated', 'suggest', 'suggests', 'suggested', 'suggesting', 'suggested', 'imply',
    'implies', 'implied', 'implying', 'implied', 'conclude', 'concludes', 'concluded',
    'concluding', 'concluded', 'report', 'reports', 'reported', 'reporting', 'reported',
    'present', 'presents', 'presented', 'presenting', 'presented', 'discuss', 'discusses',
    'discussed', 'discussing', 'discussed', 'consider', 'considers', 'considered',
    'considering', 'considered', 'review', 'reviews', 'reviewed', 'reviewing', 'reviewed',
    'describe', 'describes', 'described', 'describing', 'described', 'explain', 'explains',
    'explained', 'explaining', 'explained', 'define', 'defines', 'defined', 'defining',
    'defined', 'mean', 'means', 'meant', 'meaning', 'meant', 'refer', 'refers', 'referred',
    'referring', 'referred', 'relate', 'relates', 'related', 'relating', 'related', 'associate',
    'associates', 'associated', 'associating', 'associated', 'connect', 'connects', 'connected',
    'connecting', 'connected', 'link', 'links', 'linked', 'linking', 'linked', 'affect',
    'affects', 'affected', 'affecting', 'affected', 'effect', 'effects', 'effected',
    'effecting', 'effected', 'cause', 'causes', 'caused', 'causing', 'caused', 'result',
    'results', 'resulted', 'resulting', 'resulted', 'lead', 'leads', 'led', 'leading',
    'led', 'produce', 'produces', 'produced', 'producing', 'produced', 'create', 'creates',
    'created', 'creating', 'created', 'develop', 'develops', 'developed', 'developing',
    'developed', 'improve', 'improves', 'improved', 'improving', 'improved', 'increase',
    'increases', 'increased', 'increasing', 'increased', 'decrease', 'decreases', 'decreased',
    'decreasing', 'decreased', 'change', 'changes', 'changed', 'changing', 'changed',
    'vary', 'varies', 'varied', 'varying', 'varied', 'differ', 'differs', 'differed',
    'differing', 'differed', 'similar', 'similarly', 'different', 'differently', 'same',
    'differently', 'other', 'others', 'another', 'more', 'less', 'most', 'least', 'many',
    'much', 'few', 'little', 'several', 'some', 'any', 'all', 'none', 'every', 'each',
    'either', 'neither', 'both', 'such', 'so', 'too', 'very', 'extremely', 'highly',
    'greatly', 'strongly', 'significantly', 'remarkably', 'notably', 'particularly',
    'specifically', 'especially', 'mainly', 'largely', 'mostly', 'primarily', 'principally',
    'generally', 'usually', 'often', 'frequently', 'sometimes', 'occasionally', 'rarely',
    'seldom', 'never', 'always', 'constantly', 'continuously', 'regularly', 'periodically',
    'initially', 'finally', 'eventually', 'ultimately', 'subsequently', 'consequently',
    'therefore', 'thus', 'hence', 'accordingly', 'however', 'nevertheless', 'nonetheless',
    'though', 'although', 'even', 'even though', 'despite', 'in spite of', 'because',
    'since', 'as', 'for', 'due to', 'owing to', 'on account of', 'because of', 'if',
    'whether', 'when', 'whenever', 'where', 'wherever', 'while', 'as long as', 'until',
    'unless', 'once', 'after', 'before', 'during', 'between', 'among', 'amongst', 'through',
    'throughout', 'across', 'along', 'over', 'above', 'below', 'under', 'beneath', 'on',
    'onto', 'off', 'out', 'in', 'into', 'within', 'without', 'by', 'with', 'about',
    'around', 'near', 'beside', 'next to', 'against', 'opposite', 'towards', 'to',
    'from', 'of', 'for', 'at', 'by', 'with', 'about', 'against', 'between', 'into',
    'through', 'during', 'before', 'after', 'above', 'below', 'to', 'from', 'up',
    'down', 'in', 'out', 'on', 'off', 'over', 'under', 'again', 'further', 'then',
    'once', 'here', 'there', 'all', 'any', 'both', 'each', 'few', 'more', 'most',
    'other', 'some', 'such', 'no', 'nor', 'not', 'only', 'own', 'same', 'so', 'than',
    'too', 'very', 'can', 'will', 'just', 'don', 'should', 'now',
    
    # 用户提到的无研究意义的词汇
    'students', 'learning', 'education', 'an', 'student', 'university', 'based', 'data', 'these', 'teaching',
    'analysis', 'study', 'using', 'results', 'model', 'approach', 'method', 'system', 'design', 'implementation',
    'application', 'impact', 'performance', 'evaluation', 'development', 'framework', 'case', 'study', 'algorithm',
    'effect', 'process', 'methodology', 'theory', 'practice', 'context', 'perspective', 'field', 'area',
    'literature', 'review', 'background', 'motivation', 'contribution', 'challenge', 'problem', 'solution',
    'opportunity', 'trend', 'future', 'direction', 'implication', 'limitation', 'conclusion', 'discussion',
    'recommendation', 'suggestion', 'proposal', 'strategy', 'policy', 'measure', 'intervention', 'initiative',
    'program', 'project', 'activity', 'effort', 'practice', 'experience', 'observation', 'finding', 'discovery',
    'insight', 'understanding', 'knowledge', 'awareness', 'perception', 'attitude', 'behavior', 'action',
    'reaction', 'response', 'interaction', 'relationship', 'connection', 'association', 'correlation', 'link',
    'impact', 'influence', 'effect', 'cause', 'reason', 'factor', 'element', 'component', 'aspect', 'dimension',
    'characteristic', 'feature', 'attribute', 'property', 'quality', 'trait', 'nature', 'type', 'kind',
    'category', 'classification', 'group', 'set', 'collection', 'sample', 'population', 'cohort', 'case',
    'number', 'amount', 'quantity', 'volume', 'size', 'scale', 'extent', 'degree', 'level', 'intensity',
    'frequency', 'rate', 'speed', 'velocity', 'duration', 'time', 'period', 'phase', 'stage',
    'process', 'procedure', 'method', 'technique', 'tool', 'instrument', 'device', 'equipment',
    'system', 'platform', 'framework', 'model', 'theory', 'concept', 'idea', 'notion', 'principle',
    'law', 'rule', 'standard', 'norm', 'criterion', 'measure', 'metric', 'indicator', 'parameter',
    'variable', 'factor', 'element', 'component', 'unit', 'quantity', 'value', 'amount', 'number',
    
    # 中文常见停用词
    '研究', '分析', '影响', '发展', '关系', '作用', '方法', '结果',
    '讨论', '结论', '目的', '意义', '过程', '问题', '现状', '趋势', '建议',
    '措施', '对策', '政策', '系统', '模型', '理论', '实践', '应用', '技术',
    '方法', '手段', '工具', '平台', '机制', '功能', '结构', '特征', '性质',
    '类型', '分类', '定义', '概念', '原理', '原则', '规律', '法则', '定理',
    '定律', '现象', '事实', '数据', '信息', '知识', '认识', '理解', '解释',
    '说明', '阐述', '论述', '表达', '描述', '概括', '总结', '归纳', '演绎',
    '推理', '论证', '证明', '验证', '检验', '测试', '实验', '试验', '观察',
    '调查', '研究', '探讨', '探究', '探索', '发现', '发明', '创造', '创新',
    '发展', '进步', '提高', '改善', '改进', '完善', '优化', '升级', '更新',
    '变革', '变化', '转变', '转换', '转移', '迁移', '演变', '演化', '进化',
    '退化', '衰亡', '消亡', '消失', '灭绝', '存在', '出现', '产生', '发生',
    '形成', '构成', '组成', '结构', '构造', '组织', '系统', '体系', '体制',
    '机制', '制度', '政策', '方针', '策略', '方法', '方式', '模式', '形式',
    '格式', '样式', '类型', '种类', '类别', '分类', '分级', '层次', '等级',
    '程度', '水平', '水准', '标准', '规范', '规则', '准则', '原则', '原理',
    '理论', '学说', '观点', '看法', '认识', '理解', '意识', '思想', '思维',
    '思考', '考虑', '分析', '判断', '推理', '论证', '证明', '验证', '检验',
    '测试', '实验', '试验', '观察', '调查'
}

# 1. 读取高活跃度作者列表
high_file = r'f:\projects\nlp\NLP-Midterm\r1\csv\highly_productive_scholars.csv'
high_authors = set()
with open(high_file, 'r', encoding='utf-8-sig') as f:
    reader = csv.DictReader(f)
    for row in reader:
        high_authors.add(row['Author'].strip())

print(f"高活跃度作者数量: {len(high_authors)}")

# 2. 从04表中提取关键词
excel_file = r'f:\projects\nlp\NLP-Midterm\r1\csv\04_Diaspora_Articles_12195(1).xlsx'

# 使用read_only模式打开Excel文件
import openpyxl
workbook = openpyxl.load_workbook(excel_file, read_only=True)
sheet = workbook.active

# 找到需要的列
author_col = None
keywords_col = None
keywords_plus_col = None
row = sheet[1]  # 获取表头行
for cell in row:
    if cell.value == 'Diaspora_Author_Names':
        author_col = cell.column - 1  # 转换为0-based索引
    elif cell.value == 'Author Keywords':
        keywords_col = cell.column - 1  # 转换为0-based索引
    elif cell.value == 'Keywords Plus':
        keywords_plus_col = cell.column - 1  # 转换为0-based索引

if author_col is None or keywords_col is None:
    print("错误：未找到所需的列")
    exit(1)

# 收集关键词
all_keywords = []
high_keywords = []
low_keywords = []

processed_count = 0
batch_size = 1000
max_rows = sheet.max_row

for row in sheet.iter_rows(min_row=2, max_row=max_rows):
    processed_count += 1
    if processed_count % batch_size == 0:
        print(f"已处理 {processed_count} / {max_rows} 行")
    
    # 获取作者、关键词和摘要信息
    author_cell = row[author_col].value
    keywords_cell = row[keywords_col].value
    keywords_plus_cell = row[keywords_plus_col].value if keywords_plus_col is not None else None
    
    # 仅从关键词字段提取
    paper_keywords = []
    if keywords_cell:
        paper_keywords.extend([kw.strip().lower() for kw in str(keywords_cell).split(';') if kw.strip()])
    if keywords_plus_cell:
        paper_keywords.extend([kw.strip().lower() for kw in str(keywords_plus_cell).split(';') if kw.strip()])
    
    if author_cell and paper_keywords:
        # 处理多个作者的情况
        paper_authors = [author.strip() for author in str(author_cell).split(';') if author.strip()]
        
        # 判断论文是否属于高活跃度作者
        is_high_author_paper = any(author in high_authors for author in paper_authors)
        
        # 过滤停用词
        filtered_keywords = [kw for kw in paper_keywords if kw not in stopwords and len(kw) > 1]
        
        if filtered_keywords:
            # 添加到所有论文的关键词列表
            all_keywords.extend(filtered_keywords)
            
            # 添加到相应的作者类别列表
            if is_high_author_paper:
                high_keywords.extend(filtered_keywords)
            else:
                low_keywords.extend(filtered_keywords)

workbook.close()

print(f"已处理 {processed_count} 行")
print(f"所有论文关键词数量: {len(all_keywords)}")
print(f"高活跃度作者论文关键词数量: {len(high_keywords)}")
print(f"非高活跃度作者论文关键词数量: {len(low_keywords)}")

# 3. 计算高频关键词
def get_top_keywords(keywords, top_n=10):
    """获取前n个高频关键词"""
    word_freq = Counter(keywords)
    return word_freq.most_common(top_n)

# 获取前15个高频关键词
all_top_15 = get_top_keywords(all_keywords, 15)
high_top_15 = get_top_keywords(high_keywords, 15)
low_top_15 = get_top_keywords(low_keywords, 15)

# 4. 显示对比结果
print("\n=== 关键词对比结果 ===")

# 打印前15个高频关键词
print("\n1. 所有论文前15个高频关键词:")
for i, (keyword, freq) in enumerate(all_top_15, 1):
    print(f"   {i}. {keyword}: {freq} 次")

print("\n2. 高活跃度作者论文前15个高频关键词:")
for i, (keyword, freq) in enumerate(high_top_15, 1):
    print(f"   {i}. {keyword}: {freq} 次")

print("\n3. 非高活跃度作者论文前15个高频关键词:")
for i, (keyword, freq) in enumerate(low_top_15, 1):
    print(f"   {i}. {keyword}: {freq} 次")

# 5. 分析共同和独特关键词
all_keyword_set = set([kw for kw, _ in all_top_15])
high_keyword_set = set([kw for kw, _ in high_top_15])
low_keyword_set = set([kw for kw, _ in low_top_15])

common_keywords = high_keyword_set.intersection(low_keyword_set)
unique_high_keywords = high_keyword_set - low_keyword_set
unique_low_keywords = low_keyword_set - high_keyword_set

print("\n4. 关键词对比分析:")
print(f"   - 共同关键词数量: {len(common_keywords)}")
print(f"   - 共同关键词: {', '.join(sorted(common_keywords))}")
print(f"   - 高活跃度作者特有关键词: {', '.join(sorted(unique_high_keywords))}")
print(f"   - 非高活跃度作者特有关键词: {', '.join(sorted(unique_low_keywords))}")

# 6. 生成对比表格
print("\n5. 关键词对比表格:")
print("   {:<15} {:<10} {:<10} {:<10}".format("关键词", "所有论文", "高活跃度", "非高活跃度"))
print("   " + "="*45)

# 合并所有关键词
top_keywords = sorted(all_keyword_set.union(high_keyword_set).union(low_keyword_set))

# 创建频率映射
all_freq_map = dict(all_top_15)
high_freq_map = dict(high_top_15)
low_freq_map = dict(low_top_15)

# 打印表格
for keyword in top_keywords:
    all_freq = all_freq_map.get(keyword, 0)
    high_freq = high_freq_map.get(keyword, 0)
    low_freq = low_freq_map.get(keyword, 0)
    print("   {:<15} {:<10} {:<10} {:<10}".format(keyword, all_freq, high_freq, low_freq))

print("\n=== 分析完成！ ===")
