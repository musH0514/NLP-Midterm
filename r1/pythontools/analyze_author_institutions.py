import openpyxl
import csv
import os
import re
from collections import defaultdict, Counter
import matplotlib.pyplot as plt

print("开始分析学者机构性质...")

# 1. 定义机构类别和关键词
print("定义机构类别和关键词...")

institution_categories = {
    '高校': [
        r'University', r'College', r'School', r'Institute.*Technology',
        r'Polytechnic', r'Academy', r'Faculty', r'Center.*University',
        r'University.*Center', r'Department.*University', r'University.*Department',
        r'Graduate.*School', r'School.*Graduate', r'National.*University',
        r'Universidad', r'Université', r'Universität', r' Universität',  # 其他语言的大学
        r'College of', r'School of', r'University of', r'State University',
        r'Community College', r'Technical College', r'Junior College'
    ],
    '科研院所': [
        r'Research.*Institute', r'Institute.*Research', r'Academy of.*Science',
        r'Science.*Academy', r'National.*Laboratory', r'Laboratory.*National',
        r'Research.*Center', r'Center.*Research', r'Institute of.*',
        r'Chinese Academy of.*', r'American.*Institute', r'European.*Institute',
        r'International.*Institute', r'Research.*Foundation', r'Foundation.*Research',
        r'Scientific.*Research', r'Research.*Scientific', r'Advanced.*Research'
    ],
    '企业': [
        r'Company', r'Corporation', r'Inc\.', r'Ltd\.', r'Limited',
        r'Co\.', r'LLC', r'Pty Ltd', r'AG', r'GmbH',
        r'Industry', r'Technology.*Corp', r'Corp.*Technology',
        r'Software.*Company', r'Company.*Software', r'Engineering.*Company',
        r'Company.*Engineering', r'Group.*Company', r'Company.*Group',
        r'Enterprises', r'Enterprise', r'Business', r'Firm'
    ],
    '医疗机构': [
        r'Hospital', r'Clinic', r'Medical.*Center', r'Center.*Medical',
        r'Healthcare', r'Health.*Center', r'Center.*Health',
        r'Medicine.*Department', r'Department.*Medicine', r'Medical.*School',
        r'School.*Medicine', r'Health.*Science', r'Science.*Health',
        r'Pharmaceutical', r'Pharmacy', r'Clinical.*Research'
    ]
}

# 2. 读取unique_authors_with_subject.csv文件
print("读取学者列表...")
subject_file = r'f:\projects\nlp\NLP-Midterm\r1\csv\unique_authors_with_subject.csv'

authors = []
with open(subject_file, 'r', encoding='utf-8-sig') as f:
    reader = csv.DictReader(f)
    for row in reader:
        authors.append(row['Author'])

print(f"共读取 {len(authors)} 位学者")

# 3. 从04表中提取学者的机构信息
print("\n从04表中提取机构信息...")
excel_file = r'f:\projects\nlp\NLP-Midterm\r1\csv\04_Diaspora_Articles_12195(1).xlsx'

# 使用read_only模式打开Excel文件
workbook = openpyxl.load_workbook(excel_file, read_only=True)
sheet = workbook.active

# 找到需要的列
author_col = None
affiliation_col = None
row = sheet[1]  # 获取表头行
for cell in row:
    if cell.value == 'Diaspora_Author_Names':
        author_col = cell.column - 1  # 转换为0-based索引
    elif cell.value == 'Affiliations':
        affiliation_col = cell.column - 1  # 转换为0-based索引

if author_col is None or affiliation_col is None:
    print("错误：未找到所需的列")
    exit(1)

print(f"Diaspora_Author_Names列在索引 {author_col}")
print(f"Affiliations列在索引 {affiliation_col}")

# 收集每个学者的机构信息
author_institutions = defaultdict(list)
processed_count = 0
batch_size = 1000
max_rows = sheet.max_row

for row in sheet.iter_rows(min_row=2, max_row=max_rows):
    processed_count += 1
    if processed_count % batch_size == 0:
        print(f"已处理 {processed_count} / {max_rows} 行")
    
    # 获取作者和机构信息
    author_cell = row[author_col].value
    affiliation_cell = row[affiliation_col].value
    
    if author_cell:
        # 处理多个作者的情况
        paper_authors = [author.strip() for author in str(author_cell).split(';') if author.strip()]
        
        if affiliation_cell:
            # 处理多个机构的情况
            paper_institutions = [inst.strip() for inst in str(affiliation_cell).split(';') if inst.strip()]
            
            for author in paper_authors:
                if author in authors:  # 只处理unique_authors中的学者
                    author_institutions[author].extend(paper_institutions)

workbook.close()

print(f"已处理 {processed_count} 行")
print(f"已提取 {len(author_institutions)} 位学者的机构信息")

# 4. 为每个机构确定其类别
def classify_institution(institution):
    """将机构分类到指定类别"""
    institution_lower = institution.lower()
    
    for category, keywords in institution_categories.items():
        for keyword in keywords:
            if re.search(keyword, institution, re.IGNORECASE):
                return category
    
    return '其他'

# 5. 分析每个学者的机构类别
print("\n分析学者的机构类别...")
author_institution_categories = defaultdict(list)

for author, institutions in author_institutions.items():
    # 统计每个机构出现的次数
    institution_counts = Counter(institutions)
    # 找到出现次数最多的机构（主要机构）
    if institution_counts:
        main_institution = institution_counts.most_common(1)[0][0]
        # 分类主要机构
        main_category = classify_institution(main_institution)
        # 也获取所有机构的类别
        all_categories = list(set([classify_institution(inst) for inst in institutions]))
        
        author_institution_categories[author] = {
            'main_institution': main_institution,
            'main_category': main_category,
            'all_institutions': institutions,
            'all_categories': all_categories
        }
    else:
        author_institution_categories[author] = {
            'main_institution': '未知',
            'main_category': '未知',
            'all_institutions': [],
            'all_categories': ['未知']
        }

# 6. 生成包含机构信息的新表
print("\n生成包含机构信息的新表...")

output_file = r'f:\projects\nlp\NLP-Midterm\r1\csv\authors_with_institution_categories.csv'

with open(output_file, 'w', newline='', encoding='utf-8-sig') as f:
    writer = csv.writer(f)
    # 写入表头
    writer.writerow([
        'Author', 'Main_Institution', 'Main_Institution_Category',
        'All_Institutions', 'All_Institution_Categories'
    ])
    
    # 写入数据
    for author in authors:
        if author in author_institution_categories:
            info = author_institution_categories[author]
            main_institution = info['main_institution']
            main_category = info['main_category']
            all_institutions = '; '.join(info['all_institutions'][:10])  # 最多保留10个机构
            all_categories = '; '.join(info['all_categories'])
        else:
            main_institution = '未知'
            main_category = '未知'
            all_institutions = ''
            all_categories = '未知'
        
        writer.writerow([
            author, main_institution, main_category,
            all_institutions, all_categories
        ])

print(f"包含机构信息的新表已保存到: {output_file}")

# 7. 统计每个机构类别的人数
print("\n统计机构类别人数...")

# 按主要机构类别统计
main_category_counts = Counter()
# 按所有机构类别统计（每个学者可能属于多个类别）
all_category_counts = Counter()

for author, info in author_institution_categories.items():
    main_category_counts[info['main_category']] += 1
    for category in info['all_categories']:
        all_category_counts[category] += 1

print("\n按主要机构类别统计:")
for category, count in main_category_counts.most_common():
    print(f"{category}: {count} 人")

print("\n按所有机构类别统计:")
for category, count in all_category_counts.most_common():
    print(f"{category}: {count} 人")

# 8. 生成机构类别统计的新表
print("\n生成机构类别统计的新表...")

stats_file = r'f:\projects\nlp\NLP-Midterm\r1\csv\institution_category_stats.csv'

with open(stats_file, 'w', newline='', encoding='utf-8-sig') as f:
    writer = csv.writer(f)
    # 写入表头
    writer.writerow(['Category', 'Main_Category_Count', 'All_Categories_Count'])
    
    # 获取所有唯一类别
    all_categories = sorted(set(main_category_counts.keys()) | set(all_category_counts.keys()))
    
    # 写入数据
    for category in all_categories:
        main_count = main_category_counts.get(category, 0)
        all_count = all_category_counts.get(category, 0)
        writer.writerow([category, main_count, all_count])

print(f"机构类别统计的新表已保存到: {stats_file}")

# 9. 生成机构类别分布图表
print("\n生成机构类别分布图表...")

# 创建图表目录
output_dir = r'f:\projects\nlp\NLP-Midterm\r1\maps\background'
if not os.path.exists(output_dir):
    os.makedirs(output_dir)

# 设置中文字体支持
plt.rcParams['font.sans-serif'] = ['SimHei', 'Arial Unicode MS', 'DejaVu Sans']
plt.rcParams['axes.unicode_minus'] = False

# 图表1: 主要机构类别分布
plt.figure(figsize=(10, 6))

# 准备数据
main_categories = [category for category, _ in main_category_counts.most_common()]
main_counts = [count for _, count in main_category_counts.most_common()]
colors = ['#FF6B6B', '#4ECDC4', '#45B7D1', '#FFA07A', '#98D8C8', '#F7DC6F']

# 创建饼图
plt.pie(main_counts, labels=main_categories, colors=colors[:len(main_categories)], autopct='%1.1f%%', startangle=90)

plt.title('学者主要机构类别分布', fontsize=16)
plt.axis('equal')  # 保证饼图是圆形

# 保存图表
chart1_file = os.path.join(output_dir, 'institution_main_category_distribution.png')
plt.savefig(chart1_file, dpi=300, bbox_inches='tight')
print(f"主要机构类别分布饼图已保存到: {chart1_file}")

# 图表2: 所有机构类别分布
plt.figure(figsize=(10, 6))

# 准备数据
all_categories = [category for category, _ in all_category_counts.most_common()]
all_counts = [count for _, count in all_category_counts.most_common()]

# 创建柱状图
bars = plt.bar(all_categories, all_counts, color=colors[:len(all_categories)])

plt.title('学者所有机构类别分布', fontsize=16)
plt.xlabel('机构类别', fontsize=12)
plt.ylabel('人数', fontsize=12)
plt.xticks(rotation=45, ha='right')

# 在柱子上显示数值
for bar in bars:
    height = bar.get_height()
    plt.text(bar.get_x() + bar.get_width()/2., height + 10,
             f'{height}', ha='center', va='bottom')

# 添加网格线
plt.grid(axis='y', alpha=0.3)

# 调整布局
plt.tight_layout()

# 保存图表
chart2_file = os.path.join(output_dir, 'institution_all_categories_distribution.png')
plt.savefig(chart2_file, dpi=300, bbox_inches='tight')
print(f"所有机构类别分布柱状图已保存到: {chart2_file}")

# 10. 生成HTML版本的图表
print("\n生成HTML版本的图表...")

# 准备数据
main_category_data = []
for category, count in main_category_counts.most_common():
    main_category_data.append({
        'category': category,
        'count': count
    })

all_category_data = []
for category, count in all_category_counts.most_common():
    all_category_data.append({
        'category': category,
        'count': count
    })

# 创建HTML内容
html_content = '''<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>学者机构类别分布</title>
    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
    <style>
        body {
            font-family: Arial, sans-serif;
            margin: 0;
            padding: 20px;
            background-color: #f5f5f5;
        }
        .container {
            max-width: 1200px;
            margin: 0 auto;
            background-color: white;
            padding: 30px;
            border-radius: 10px;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        }
        h1 {
            text-align: center;
            color: #333;
        }
        h2 {
            color: #555;
            border-bottom: 2px solid #4ECDC4;
            padding-bottom: 10px;
        }
        .chart-container {
            margin: 40px 0;
            height: 400px;
        }
        table {
            width: 100%;
            border-collapse: collapse;
            margin: 20px 0;
        }
        th, td {
            padding: 12px;
            text-align: left;
            border-bottom: 1px solid #ddd;
        }
        th {
            background-color: #4ECDC4;
            color: white;
        }
        tr:hover {
            background-color: #f5f5f5;
        }
    </style>
</head>
<body>
    <div class="container">
        <h1>学者机构类别分布</h1>
        
        <h2>1. 主要机构类别分布</h2>
        <div class="chart-container">
            <canvas id="mainCategoryChart"></canvas>
        </div>
        
        <h2>2. 所有机构类别分布</h2>
        <div class="chart-container">
            <canvas id="allCategoryChart"></canvas>
        </div>
        
        <h2>3. 机构类别统计数据</h2>
        <table>
            <tr>
                <th>机构类别</th>
                <th>主要机构类别人数</th>
                <th>所有机构类别人数</th>
            </tr>
            {table_data}
        </table>
    </div>
    
    <script>
        // 主要机构类别饼图
        const mainCtx = document.getElementById('mainCategoryChart').getContext('2d');
        const mainCategoryChart = new Chart(mainCtx, {{
            type: 'pie',
            data: {{
                labels: {main_labels},
                datasets: [{{
                    data: {main_data},
                    backgroundColor: ['#FF6B6B', '#4ECDC4', '#45B7D1', '#FFA07A', '#98D8C8', '#F7DC6F'],
                    borderWidth: 1
                }}]
            }},
            options: {{
                responsive: true,
                maintainAspectRatio: false,
                plugins: {{
                    title: {{
                        display: true,
                        text: '学者主要机构类别分布',
                        font: {{
                            size: 16
                        }}
                    }},
                    legend: {{
                        position: 'bottom'
                    }},
                    tooltip: {{
                        callbacks: {{
                            label: function(context) {{
                                const label = context.label || '';
                                const value = context.parsed || 0;
                                const total = context.dataset.data.reduce((a, b) => a + b, 0);
                                const percentage = ((value / total) * 100).toFixed(1);
                                return `${{label}}: ${{value}} (${{percentage}}%)`;
                            }}
                        }}
                    }}
                }}
            }}
        }});
        
        // 所有机构类别柱状图
        const allCtx = document.getElementById('allCategoryChart').getContext('2d');
        const allCategoryChart = new Chart(allCtx, {{
            type: 'bar',
            data: {{
                labels: {all_labels},
                datasets: [{{
                    label: '人数',
                    data: {all_data},
                    backgroundColor: ['#FF6B6B', '#4ECDC4', '#45B7D1', '#FFA07A', '#98D8C8', '#F7DC6F'],
                    borderWidth: 1
                }}]
            }},
            options: {{
                responsive: true,
                maintainAspectRatio: false,
                plugins: {{
                    title: {{
                        display: true,
                        text: '学者所有机构类别分布',
                        font: {{
                            size: 16
                        }}
                    }},
                    legend: {{
                        display: false
                    }},
                    tooltip: {{
                        callbacks: {{
                            label: function(context) {{
                                return '人数: ' + context.parsed.y;
                            }}
                        }}
                    }}
                }},
                scales: {{
                    y: {{
                        beginAtZero: true,
                        ticks: {{
                            stepSize: 500
                        }}
                    }}
                }}
            }}
        }});
    </script>
</body>
</html>'''

# 生成表格数据
table_rows = []
for category in all_categories:
    main_count = main_category_counts.get(category, 0)
    all_count = all_category_counts.get(category, 0)
    table_rows.append(f"<tr><td>{category}</td><td>{main_count}</td><td>{all_count}</td></tr>")
table_data = '\n            '.join(table_rows)

# 准备图表数据
main_labels = [category for category, _ in main_category_counts.most_common()]
main_data = [count for _, count in main_category_counts.most_common()]
all_labels = [category for category, _ in all_category_counts.most_common()]
all_data = [count for _, count in all_category_counts.most_common()]

# 替换HTML中的占位符
html_content = html_content.replace('{table_data}', table_data)
html_content = html_content.replace('{main_labels}', str(main_labels).replace("'", '"'))
html_content = html_content.replace('{main_data}', str(main_data))
html_content = html_content.replace('{all_labels}', str(all_labels).replace("'", '"'))
html_content = html_content.replace('{all_data}', str(all_data))

# 保存HTML文件
html_file = os.path.join(output_dir, 'institution_category_distribution.html')
with open(html_file, 'w', encoding='utf-8') as f:
    f.write(html_content)

print(f"HTML版本的图表已保存到: {html_file}")

print("\n=== 分析完成！ ===")
print("\n生成的文件:")
print(f"1. {output_file} - 包含机构信息的学者表")
print(f"2. {stats_file} - 机构类别统计数据")
print(f"3. {chart1_file} - 主要机构类别分布饼图")
print(f"4. {chart2_file} - 所有机构类别分布柱状图")
print(f"5. {html_file} - HTML版本的机构类别分布图表")
