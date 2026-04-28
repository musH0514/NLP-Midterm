import openpyxl
import os
from collections import Counter

# 读取Excel文件，只获取必要的列
excel_file = r'f:\projects\nlp\NLP-Midterm\r1\csv\04_Diaspora_Articles_12195(1).xlsx'

print(f"正在读取Excel文件: {excel_file}")
print(f"文件是否存在: {os.path.exists(excel_file)}")

try:
    # 使用read_only模式打开文件
    workbook = openpyxl.load_workbook(excel_file, read_only=True)
    sheet = workbook.active
    
    # 获取表头
    headers = []
    for cell in sheet[1]:
        headers.append(cell.value)
    
    # 找到需要的列索引
    author_col_idx = headers.index('Diaspora_Author_Names') if 'Diaspora_Author_Names' in headers else -1
    wos_col_idx = headers.index('WoS Categories') if 'WoS Categories' in headers else -1
    research_ids_idx = headers.index('Researcher Ids') if 'Researcher Ids' in headers else -1
    orcids_idx = headers.index('ORCIDs') if 'ORCIDs' in headers else -1
    title_idx = headers.index('Article Title') if 'Article Title' in headers else -1
    year_idx = headers.index('Publication Year') if 'Publication Year' in headers else -1
    
    print(f"\n找到的列索引:")
    print(f"  Diaspora_Author_Names: {author_col_idx}")
    print(f"  WoS Categories: {wos_col_idx}")
    print(f"  Researcher Ids: {research_ids_idx}")
    print(f"  ORCIDs: {orcids_idx}")
    
    # 查找作者为"An, Brian P."的所有记录
    author_name = "An, Brian P."
    author_records = []
    
    print(f"\n开始搜索包含 '{author_name}' 的所有记录...")
    
    # 逐行读取数据
    total_rows = sheet.max_row
    for row_num, row in enumerate(sheet.iter_rows(min_row=2, max_row=total_rows), start=2):
        # 检查作者列
        if author_col_idx != -1 and author_col_idx < len(row) and row[author_col_idx].value:
            if author_name in str(row[author_col_idx].value):
                # 收集记录信息
                record = {
                    'row_num': row_num
                }
                
                # 添加标题
                if title_idx != -1 and title_idx < len(row):
                    record['title'] = row[title_idx].value
                
                # 添加WoS分类
                if wos_col_idx != -1 and wos_col_idx < len(row):
                    record['wos_categories'] = row[wos_col_idx].value
                
                # 添加年份
                if year_idx != -1 and year_idx < len(row):
                    record['year'] = row[year_idx].value
                
                # 添加Researcher Ids
                if research_ids_idx != -1 and research_ids_idx < len(row):
                    record['researcher_ids'] = row[research_ids_idx].value
                
                # 添加ORCIDs
                if orcids_idx != -1 and orcids_idx < len(row):
                    record['orcids'] = row[orcids_idx].value
                
                author_records.append(record)
        
        # 每处理1000行显示进度
        if row_num % 1000 == 0:
            print(f"  已处理 {row_num}/{total_rows} 行...")
    
    print(f"搜索完成！找到 {len(author_records)} 条包含 '{author_name}' 的记录")
    
    # 分析结果
    if author_records:
        print(f"\n记录详情:")
        for i, record in enumerate(author_records, 1):
            print(f"\n记录 {i} (Excel行号: {record['row_num']}):")
            if 'title' in record:
                print(f"  标题: {record['title']}")
            if 'year' in record:
                print(f"  年份: {record['year']}")
            if 'wos_categories' in record:
                print(f"  WoS分类: {record['wos_categories']}")
            if 'researcher_ids' in record and record['researcher_ids']:
                print(f"  Researcher Ids: {record['researcher_ids']}")
            if 'orcids' in record and record['orcids']:
                print(f"  ORCIDs: {record['orcids']}")
        
        # 检查Researcher Ids和ORCIDs是否有多个值
        researcher_ids = set()
        orcids = set()
        
        for record in author_records:
            if 'researcher_ids' in record and record['researcher_ids']:
                researcher_ids.add(str(record['researcher_ids']))
            if 'orcids' in record and record['orcids']:
                orcids.add(str(record['orcids']))
        
        print(f"\n作者标识分析:")
        print(f"  不同的Researcher Ids数量: {len(researcher_ids)}")
        if researcher_ids:
            print(f"    值: {researcher_ids}")
        print(f"  不同的ORCIDs数量: {len(orcids)}")
        if orcids:
            print(f"    值: {orcids}")
        
        # 分析学科分布
        print(f"\n学科分布分析:")
        all_categories = []
        for record in author_records:
            if 'wos_categories' in record and record['wos_categories']:
                wos_cats = str(record['wos_categories']).split('; ')
                all_categories.extend(wos_cats)
        
        category_counts = Counter(all_categories)
        print(f"  共涉及 {len(category_counts)} 个学科分类")
        for cat, count in sorted(category_counts.items(), key=lambda x: x[1], reverse=True):
            if cat:
                print(f"    {cat}: {count} 篇")
    
    print(f"\n结论:")
    if len(author_records) == 16:
        print(f"  找到了与unique_authors_with_subject.csv中一致的16篇论文")
    else:
        print(f"  找到 {len(author_records)} 篇论文，与unique_authors_with_subject.csv中的16篇不一致")
    
    if len(researcher_ids) > 1 or len(orcids) > 1:
        print(f"  ⚠️  发现多个不同的作者标识，说明 '{author_name}' 很可能是多个同名不同作者的集合")
    else:
        print(f"  所有论文都有相同的作者标识，说明 '{author_name}' 是同一个作者发表了跨领域的论文")
    
    workbook.close()
    
except Exception as e:
    print(f"处理过程中出错: {e}")
    import traceback
    traceback.print_exc()
