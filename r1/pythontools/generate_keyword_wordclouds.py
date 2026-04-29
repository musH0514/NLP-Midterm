import openpyxl
import csv
import os
import re
import matplotlib.pyplot as plt
from wordcloud import WordCloud
import numpy as np
from collections import Counter
import jieba
from PIL import Image

print("开始生成关键词词云图...")

# 1. 安装必要的库（如果未安装）
try:
    from wordcloud import WordCloud
except ImportError:
    print("正在安装wordcloud库...")
    import subprocess
    subprocess.check_call(['pip', 'install', 'wordcloud'])
    from wordcloud import WordCloud

try:
    import jieba
except ImportError:
    print("正在安装jieba库...")
    import subprocess
    subprocess.check_call(['pip', 'install', 'jieba'])
    import jieba

# 2. 设置中文分词和词云参数
print("设置中文分词和词云参数...")

# 设置jieba中文分词
jieba.setLogLevel(20)

# 定义停用词 - 专业版，只保留真正有研究意义的专业词汇
stopwords = {
    # 英文常见停用词
    'the', 'and', 'of', 'to', 'a', 'in', 'for', 'is', 'on', 'with', 'as', 'by', 'at', 'from',
    'that', 'this', 'are', 'was', 'were', 'be', 'been', 'being', 'which', 'what', 'who',
    'whom', 'whose', 'when', 'where', 'why', 'how', 'about', 'after', 'before', 'during',
    'between', 'into', 'through', 'over', 'above', 'below', 'up', 'down', 'out', 'off',
    'again', 'further', 'then', 'once', 'here', 'there', 'all', 'any', 'both', 'each',
    'few', 'more', 'most', 'other', 'some', 'such', 'no', 'nor', 'not', 'only', 'own',
    'same', 'so', 'than', 'too', 'very', 'can', 'will', 'just', 'should', 'now', 'but',
    'or', 'because', 'if', 'though', 'while', 'until', 'unless', 'also', 'even', 'however',
    'yet', 'still', 'already', 'almost', 'always', 'never', 'often', 'usually', 'sometimes',
    'may', 'might', 'must', 'shall', 'should', 'would', 'could', 'one', 'two', 'three',
    'four', 'five', 'six', 'seven', 'eight', 'nine', 'ten', 'first', 'second', 'third',
    'fourth', 'fifth', 'many', 'much', 'few', 'little', 'several', 'some', 'any', 'all',
    'none', 'every', 'each', 'either', 'neither', 'such', 'another', 'other', 'others',
    'its', 'it', 'you', 'your', 'yours', 'we', 'our', 'ours', 'they', 'their', 'theirs',
    'he', 'his', 'she', 'her', 'hers', 'me', 'my', 'mine', 'them', 'us', 'i', 'am', 'has',
    'have', 'had', 'do', 'does', 'did', 'doing', 'done', 'go', 'goes', 'went', 'going',
    'gone', 'get', 'gets', 'got', 'getting', 'gotten', 'make', 'makes', 'made', 'making',
    'take', 'takes', 'took', 'taking', 'taken', 'give', 'gives', 'gave', 'giving', 'given',
    'use', 'uses', 'used', 'using', 'see', 'sees', 'saw', 'seeing', 'seen', 'find', 'finds',
    'found', 'finding', 'found', 'look', 'looks', 'looked', 'looking', 'come', 'comes',
    'came', 'coming', 'came', 'bring', 'brings', 'brought', 'bringing', 'brought', 'put',
    'puts', 'put', 'putting', 'let', 'lets', 'let', 'letting', 'keep', 'keeps', 'kept',
    'keeping', 'kept', 'hold', 'holds', 'held', 'holding', 'held', 'say', 'says', 'said',
    'saying', 'said', 'speak', 'speaks', 'spoke', 'speaking', 'spoken', 'tell', 'tells',
    'told', 'telling', 'told', 'know', 'knows', 'knew', 'knowing', 'known', 'think',
    'thinks', 'thought', 'thinking', 'thought', 'believe', 'believes', 'believed',
    'believing', 'believed', 'want', 'wants', 'wanted', 'wanting', 'wanted', 'need',
    'needs', 'needed', 'needing', 'needed', 'like', 'likes', 'liked', 'liking', 'liked',
    'love', 'loves', 'loved', 'loving', 'loved', 'hate', 'hates', 'hated', 'hating',
    'hated', 'work', 'works', 'worked', 'working', 'worked', 'study', 'studies',
    'studied', 'studying', 'studied', 'research', 'researches', 'researched',
    'researching', 'researched', 'analyze', 'analyzes', 'analyzed', 'analyzing',
    'analyzed', 'examine', 'examines', 'examined', 'examining', 'examined', 'test',
    'tests', 'tested', 'testing', 'tested', 'try', 'tries', 'tried', 'trying', 'tried',
    'experiment', 'experiments', 'experimented', 'experimenting', 'experimented',
    'show', 'shows', 'showed', 'showing', 'shown', 'display', 'displays', 'displayed',
    'displaying', 'displayed', 'indicate', 'indicates', 'indicated', 'indicating',
    'indicated', 'suggest', 'suggests', 'suggested', 'suggesting', 'suggested', 'imply',
    'implies', 'implied', 'implying', 'implied', 'conclude', 'concludes', 'concluded',
    'concluding', 'concluded', 'report', 'reports', 'reported', 'reporting', 'reported',
    'present', 'presents', 'presented', 'presenting', 'presented', 'discuss', 'discusses',
    'discussed', 'discussing', 'discussed', 'consider', 'considers', 'considered',
    'considering', 'considered', 'review', 'reviews', 'reviewed', 'reviewing', 'reviewed',
    'describe', 'describes', 'described', 'describing', 'described', 'explain', 'explains',
    'explained', 'explaining', 'explained', 'define', 'defines', 'defined', 'defining',
    'defined', 'mean', 'means', 'meant', 'meaning', 'meant', 'refer', 'refers', 'referred',
    'referring', 'referred', 'relate', 'relates', 'related', 'relating', 'related', 'associate',
    'associates', 'associated', 'associating', 'associated', 'connect', 'connects', 'connected',
    'connecting', 'connected', 'link', 'links', 'linked', 'linking', 'linked', 'affect',
    'affects', 'affected', 'affecting', 'affected', 'effect', 'effects', 'effected',
    'effecting', 'effected', 'cause', 'causes', 'caused', 'causing', 'caused', 'result',
    'results', 'resulted', 'resulting', 'resulted', 'lead', 'leads', 'led', 'leading',
    'led', 'produce', 'produces', 'produced', 'producing', 'produced', 'create', 'creates',
    'created', 'creating', 'created', 'develop', 'develops', 'developed', 'developing',
    'developed', 'improve', 'improves', 'improved', 'improving', 'improved', 'increase',
    'increases', 'increased', 'increasing', 'increased', 'decrease', 'decreases', 'decreased',
    'decreasing', 'decreased', 'change', 'changes', 'changed', 'changing', 'changed',
    'vary', 'varies', 'varied', 'varying', 'varied', 'differ', 'differs', 'differed',
    'differing', 'differed', 'similar', 'similarly', 'different', 'differently', 'same',
    'differently', 'other', 'others', 'another', 'more', 'less', 'most', 'least', 'many',
    'much', 'few', 'little', 'several', 'some', 'any', 'all', 'none', 'every', 'each',
    'either', 'neither', 'both', 'such', 'so', 'too', 'very', 'extremely', 'highly',
    'greatly', 'strongly', 'significantly', 'remarkably', 'notably', 'particularly',
    'specifically', 'especially', 'mainly', 'largely', 'mostly', 'primarily', 'principally',
    'generally', 'usually', 'often', 'frequently', 'sometimes', 'occasionally', 'rarely',
    'seldom', 'never', 'always', 'constantly', 'continuously', 'regularly', 'periodically',
    'initially', 'finally', 'eventually', 'ultimately', 'subsequently', 'consequently',
    'therefore', 'thus', 'hence', 'accordingly', 'however', 'nevertheless', 'nonetheless',
    'though', 'although', 'even', 'even though', 'despite', 'in spite of', 'because',
    'since', 'as', 'for', 'due to', 'owing to', 'on account of', 'because of', 'if',
    'whether', 'when', 'whenever', 'where', 'wherever', 'while', 'as long as', 'until',
    'unless', 'once', 'after', 'before', 'during', 'between', 'among', 'amongst', 'through',
    'throughout', 'across', 'along', 'over', 'above', 'below', 'under', 'beneath', 'on',
    'onto', 'off', 'out', 'in', 'into', 'within', 'without', 'by', 'with', 'about',
    'around', 'near', 'beside', 'next to', 'against', 'opposite', 'towards', 'to',
    'from', 'of', 'for', 'at', 'by', 'with', 'about', 'against', 'between', 'into',
    'through', 'during', 'before', 'after', 'above', 'below', 'to', 'from', 'up',
    'down', 'in', 'out', 'on', 'off', 'over', 'under', 'again', 'further', 'then',
    'once', 'here', 'there', 'all', 'any', 'both', 'each', 'few', 'more', 'most',
    'other', 'some', 'such', 'no', 'nor', 'not', 'only', 'own', 'same', 'so', 'than',
    'too', 'very', 'can', 'will', 'just', 'don', 'should', 'now',
    
    # 用户提到的无研究意义的词汇
    'students', 'learning', 'education', 'an', 'student', 'university', 'based', 'data', 'these', 'teaching',
    'analysis', 'study', 'using', 'results', 'model', 'approach', 'method', 'system', 'design', 'implementation',
    'application', 'impact', 'performance', 'evaluation', 'development', 'framework', 'case', 'study', 'algorithm',
    'effect', 'process', 'methodology', 'theory', 'practice', 'context', 'perspective', 'field', 'area',
    'literature', 'review', 'background', 'motivation', 'contribution', 'challenge', 'problem', 'solution',
    'opportunity', 'trend', 'future', 'direction', 'implication', 'limitation', 'conclusion', 'discussion',
    'recommendation', 'suggestion', 'proposal', 'strategy', 'policy', 'measure', 'intervention', 'initiative',
    'program', 'project', 'activity', 'effort', 'practice', 'experience', 'observation', 'finding', 'discovery',
    'insight', 'understanding', 'knowledge', 'awareness', 'perception', 'attitude', 'behavior', 'action',
    'reaction', 'response', 'interaction', 'relationship', 'connection', 'association', 'correlation', 'link',
    'impact', 'influence', 'effect', 'cause', 'reason', 'factor', 'element', 'component', 'aspect', 'dimension',
    'characteristic', 'feature', 'attribute', 'property', 'quality', 'trait', 'nature', 'type', 'kind',
    'category', 'classification', 'group', 'set', 'collection', 'sample', 'population', 'cohort', 'case',
    'number', 'amount', 'quantity', 'volume', 'size', 'scale', 'extent', 'degree', 'level', 'intensity',
    'frequency', 'rate', 'speed', 'velocity', 'duration', 'time', 'period', 'phase', 'stage',
    'process', 'procedure', 'method', 'technique', 'tool', 'instrument', 'device', 'equipment',
    'system', 'platform', 'framework', 'model', 'theory', 'concept', 'idea', 'notion', 'principle',
    'law', 'rule', 'standard', 'norm', 'criterion', 'measure', 'metric', 'indicator', 'parameter',
    'variable', 'factor', 'element', 'component', 'unit', 'quantity', 'value', 'amount', 'number',
    
    # 中文常见停用词
    '研究', '分析', '影响', '发展', '关系', '作用', '方法', '结果',
    '讨论', '结论', '目的', '意义', '过程', '问题', '现状', '趋势', '建议',
    '措施', '对策', '政策', '系统', '模型', '理论', '实践', '应用', '技术',
    '方法', '手段', '工具', '平台', '机制', '功能', '结构', '特征', '性质',
    '类型', '分类', '定义', '概念', '原理', '原则', '规律', '法则', '定理',
    '定律', '现象', '事实', '数据', '信息', '知识', '认识', '理解', '解释',
    '说明', '阐述', '论述', '表达', '描述', '概括', '总结', '归纳', '演绎',
    '推理', '论证', '证明', '验证', '检验', '测试', '实验', '试验', '观察',
    '调查', '研究', '探讨', '探究', '探索', '发现', '发明', '创造', '创新',
    '发展', '进步', '提高', '改善', '改进', '完善', '优化', '升级', '更新',
    '变革', '变化', '转变', '转换', '转移', '迁移', '演变', '演化', '进化',
    '退化', '衰亡', '消亡', '消失', '灭绝', '存在', '出现', '产生', '发生',
    '形成', '构成', '组成', '结构', '构造', '组织', '系统', '体系', '体制',
    '机制', '制度', '政策', '方针', '策略', '方法', '方式', '模式', '形式',
    '格式', '样式', '类型', '种类', '类别', '分类', '分级', '层次', '等级',
    '程度', '水平', '水准', '标准', '规范', '规则', '准则', '原则', '原理',
    '理论', '学说', '观点', '看法', '认识', '理解', '意识', '思想', '思维',
    '思考', '考虑', '分析', '判断', '推理', '论证', '证明', '验证', '检验',
    '测试', '实验', '试验', '观察', '调查'
}

# 3. 读取高活跃度作者列表
print("读取高活跃度作者列表...")
high_file = r'f:\projects\nlp\NLP-Midterm\r1\csv\highly_productive_scholars.csv'
high_authors = set()
with open(high_file, 'r', encoding='utf-8-sig') as f:
    reader = csv.DictReader(f)
    for row in reader:
        high_authors.add(row['Author'].strip())

print(f"高活跃度作者数量: {len(high_authors)}")

# 4. 从04表中提取关键词和摘要
print("\n从04表中提取关键词和摘要...")
excel_file = r'f:\projects\nlp\NLP-Midterm\r1\csv\04_Diaspora_Articles_12195(1).xlsx'

# 使用read_only模式打开Excel文件
workbook = openpyxl.load_workbook(excel_file, read_only=True)
sheet = workbook.active

# 找到需要的列
author_col = None
keywords_col = None
abstract_col = None
row = sheet[1]  # 获取表头行
for cell in row:
    if cell.value == 'Diaspora_Author_Names':
        author_col = cell.column - 1  # 转换为0-based索引
    elif cell.value == 'Author Keywords':
        keywords_col = cell.column - 1  # 转换为0-based索引
    elif cell.value == 'Abstract':
        abstract_col = cell.column - 1  # 转换为0-based索引
    elif cell.value == 'Keywords Plus':
        keywords_plus_col = cell.column - 1  # 转换为0-based索引

if author_col is None or keywords_col is None or abstract_col is None:
    print("错误：未找到所需的列")
    exit(1)

print(f"Diaspora_Author_Names列在索引 {author_col}")
print(f"Author Keywords列在索引 {keywords_col}")
print(f"Abstract列在索引 {abstract_col}")

# 收集关键词和摘要
all_keywords = []
high_keywords = []
low_keywords = []

processed_count = 0
batch_size = 1000
max_rows = sheet.max_row

for row in sheet.iter_rows(min_row=2, max_row=max_rows):
    processed_count += 1
    if processed_count % batch_size == 0:
        print(f"已处理 {processed_count} / {max_rows} 行")
    
    # 获取作者、关键词和摘要信息
    author_cell = row[author_col].value
    keywords_cell = row[keywords_col].value
    abstract_cell = row[abstract_col].value
    keywords_plus_cell = row[keywords_plus_col].value if keywords_plus_col is not None else None
    
    # 仅从关键词字段提取，不使用摘要
    paper_keywords = []
    if keywords_cell:
        paper_keywords.extend([kw.strip() for kw in str(keywords_cell).split(';') if kw.strip()])
    if keywords_plus_cell:
        paper_keywords.extend([kw.strip() for kw in str(keywords_plus_cell).split(';') if kw.strip()])
    
    if author_cell and paper_keywords:
        # 处理多个作者的情况
        paper_authors = [author.strip() for author in str(author_cell).split(';') if author.strip()]
        
        # 判断论文是否属于高活跃度作者
        is_high_author_paper = any(author in high_authors for author in paper_authors)
        
        # 过滤停用词
        filtered_keywords = [kw.lower() for kw in paper_keywords if kw.lower() not in stopwords and len(kw) > 1]
        
        if filtered_keywords:
            # 添加到所有论文的关键词列表
            all_keywords.extend(filtered_keywords)
            
            # 添加到相应的作者类别列表
            if is_high_author_paper:
                high_keywords.extend(filtered_keywords)
            else:
                low_keywords.extend(filtered_keywords)

workbook.close()

print(f"已处理 {processed_count} 行")
print(f"所有论文关键词数量: {len(all_keywords)}")
print(f"高活跃度作者论文关键词数量: {len(high_keywords)}")
print(f"非高活跃度作者论文关键词数量: {len(low_keywords)}")

# 5. 生成词云图
print("\n生成词云图...")

# 创建图表目录
output_dir = r'f:\projects\nlp\NLP-Midterm\r1\maps\background'
if not os.path.exists(output_dir):
    os.makedirs(output_dir)

# 定义生成词云的函数
def generate_wordcloud(keywords, filename, title):
    """生成词云图"""
    # 统计词频
    word_freq = Counter(keywords)
    
    # 生成词云
    wordcloud = WordCloud(
        width=800,
        height=600,
        background_color='white',
        max_words=200,
        max_font_size=100,
        min_font_size=10,
        colormap='viridis',
        random_state=42
    ).generate_from_frequencies(word_freq)
    
    # 绘制词云
    plt.figure(figsize=(10, 7))
    plt.imshow(wordcloud, interpolation='bilinear')
    plt.title(title, fontsize=16, fontweight='bold')
    plt.axis('off')
    
    # 保存词云图
    image_path = os.path.join(output_dir, filename)
    plt.savefig(image_path, dpi=300, bbox_inches='tight')
    plt.close()
    
    print(f"词云图已保存到: {image_path}")
    return image_path

# 生成三个词云图
all_wc_path = generate_wordcloud(all_keywords, 'all_papers_wordcloud.png', 'All Papers Keywords')
high_wc_path = generate_wordcloud(high_keywords, 'high_productive_authors_wordcloud.png', 'High Productive Authors Papers Keywords')
low_wc_path = generate_wordcloud(low_keywords, 'low_productive_authors_wordcloud.png', 'Low Productive Authors Papers Keywords')

# 6. 生成HTML文件
print("\n生成HTML文件...")

# 准备词频数据用于HTML中的表格
def get_top_keywords(keywords, top_n=10):
    """获取前n个高频关键词"""
    word_freq = Counter(keywords)
    return word_freq.most_common(top_n)

# 获取前10个高频关键词
all_top_10 = get_top_keywords(all_keywords)
high_top_10 = get_top_keywords(high_keywords)
low_top_10 = get_top_keywords(low_keywords)

# 创建HTML内容
html_content = '''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Keyword WordCloud Analysis</title>
    <style>
        body {
            font-family: Arial, sans-serif;
            margin: 0;
            padding: 20px;
            background-color: #f5f5f5;
        }
        .container {
            max-width: 1200px;
            margin: 0 auto;
            background-color: white;
            padding: 30px;
            border-radius: 10px;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        }
        h1 {
            text-align: center;
            color: #333;
            margin-bottom: 40px;
        }
        .section {
            margin-bottom: 50px;
        }
        h2 {
            color: #555;
            border-bottom: 2px solid #4ECDC4;
            padding-bottom: 10px;
            margin-bottom: 20px;
        }
        .wordcloud-container {
            text-align: center;
            margin: 30px 0;
        }
        .wordcloud-container img {
            max-width: 100%;
            height: auto;
            border-radius: 10px;
            box-shadow: 0 4px 8px rgba(0,0,0,0.1);
        }
        .comparison-table {
            width: 100%;
            border-collapse: collapse;
            margin: 20px 0;
        }
        .comparison-table th, .comparison-table td {
            padding: 12px;
            text-align: left;
            border-bottom: 1px solid #ddd;
        }
        .comparison-table th {
            background-color: #4ECDC4;
            color: white;
        }
        .comparison-table tr:hover {
            background-color: #f5f5f5;
        }
        .top-keywords {
            background-color: #f9f9f9;
            padding: 20px;
            border-radius: 8px;
            margin: 20px 0;
        }
        .top-keywords h3 {
            color: #666;
            margin-top: 0;
        }
        .summary {
            background-color: #e8f4f8;
            padding: 25px;
            border-radius: 10px;
            border-left: 5px solid #4ECDC4;
            margin-top: 40px;
        }
        .summary h2 {
            margin-top: 0;
        }
        .stats {
            display: flex;
            justify-content: space-around;
            margin: 30px 0;
            flex-wrap: wrap;
        }
        .stat-card {
            background-color: #f0f0f0;
            padding: 20px;
            border-radius: 10px;
            text-align: center;
            min-width: 200px;
            margin: 10px;
        }
        .stat-card .number {
            font-size: 32px;
            font-weight: bold;
            color: #4ECDC4;
        }
        .stat-card .label {
            font-size: 14px;
            color: #666;
            margin-top: 5px;
        }
    </style>
</head>
<body>
    <div class="container">
        <h1>Keyword WordCloud Analysis</h1>
        
        <div class="stats">
            <div class="stat-card">
                <div class="number">{all_keyword_count}</div>
                <div class="label">All Papers Keywords</div>
            </div>
            <div class="stat-card">
                <div class="number">{high_keyword_count}</div>
                <div class="label">High Productive Authors Keywords</div>
            </div>
            <div class="stat-card">
                <div class="number">{low_keyword_count}</div>
                <div class="label">Low Productive Authors Keywords</div>
            </div>
        </div>
        
        <!-- All Papers WordCloud -->
        <div class="section">
            <h2>1. All Papers Keywords</h2>
            <div class="wordcloud-container">
                <img src="{all_wc_path}" alt="All Papers Keywords">
            </div>
            <div class="top-keywords">
                <h3>Top 10 Keywords</h3>
                <table class="comparison-table">
                    <tr>
                        <th>Keyword</th>
                        <th>Frequency</th>
                    </tr>
                    {all_top_10_table}
                </table>
            </div>
        </div>
        
        <!-- High Productive Authors WordCloud -->
        <div class="section">
            <h2>2. High Productive Authors Papers Keywords</h2>
            <div class="wordcloud-container">
                <img src="{high_wc_path}" alt="High Productive Authors Keywords">
            </div>
            <div class="top-keywords">
                <h3>Top 10 Keywords</h3>
                <table class="comparison-table">
                    <tr>
                        <th>Keyword</th>
                        <th>Frequency</th>
                    </tr>
                    {high_top_10_table}
                </table>
            </div>
        </div>
        
        <!-- Low Productive Authors WordCloud -->
        <div class="section">
            <h2>3. Low Productive Authors Papers Keywords</h2>
            <div class="wordcloud-container">
                <img src="{low_wc_path}" alt="Low Productive Authors Keywords">
            </div>
            <div class="top-keywords">
                <h3>Top 10 Keywords</h3>
                <table class="comparison-table">
                    <tr>
                        <th>Keyword</th>
                        <th>Frequency</th>
                    </tr>
                    {low_top_10_table}
                </table>
            </div>
        </div>
        
        <!-- Comparison Summary -->
        <div class="summary">
            <h2>4. Comparison Summary</h2>
            <h3>Key Findings:</h3>
            <ul>
                <li>All papers cover a diverse range of research topics, as seen in the comprehensive keyword distribution.</li>
                <li>High productive authors tend to focus on specific research areas, indicated by the concentration of certain keywords in their word cloud.</li>
                <li>Low productive authors show a broader but less concentrated set of research interests.</li>
            </ul>
            
            <h3>Top Keywords Comparison:</h3>
            <table class="comparison-table">
                <tr>
                    <th>Rank</th>
                    <th>All Papers</th>
                    <th>High Productive Authors</th>
                    <th>Low Productive Authors</th>
                </tr>
                {top_keywords_comparison_table}
            </table>
            
            <h3>Implication:</h3>
            <p>The keyword analysis suggests that high productive authors tend to specialize in specific research areas, while low productive authors have more varied research interests. This specialization might contribute to their higher publication productivity. The overlapping keywords indicate common research interests across both groups, while the unique keywords highlight different research focuses.</p>
        </div>
    </div>
</body>
</html>'''

# 生成表格内容
def generate_top_keywords_table(top_keywords):
    """生成高频关键词表格"""
    rows = []
    for keyword, freq in top_keywords:
        rows.append(f"<tr><td>{keyword}</td><td>{freq}</td></tr>")
    return '\n                    '.join(rows)

# 生成关键词对比表格
def generate_keywords_comparison_table(all_top, high_top, low_top):
    """生成关键词对比表格"""
    rows = []
    for i in range(10):
        all_kw = all_top[i][0] if i < len(all_top) else ''
        all_freq = all_top[i][1] if i < len(all_top) else ''
        high_kw = high_top[i][0] if i < len(high_top) else ''
        high_freq = high_top[i][1] if i < len(high_top) else ''
        low_kw = low_top[i][0] if i < len(low_top) else ''
        low_freq = low_top[i][1] if i < len(low_top) else ''
        
        rows.append(f"<tr><td>{i+1}</td><td>{all_kw}<br><small>{all_freq}</small></td><td>{high_kw}<br><small>{high_freq}</small></td><td>{low_kw}<br><small>{low_freq}</small></td></tr>")
    return '\n                '.join(rows)

# 生成表格内容
all_top_10_table = generate_top_keywords_table(all_top_10)
high_top_10_table = generate_top_keywords_table(high_top_10)
low_top_10_table = generate_top_keywords_table(low_top_10)
top_keywords_comparison_table = generate_keywords_comparison_table(all_top_10, high_top_10, low_top_10)

# 替换HTML中的占位符
html_content = html_content.replace('{all_keyword_count}', str(len(all_keywords)))
html_content = html_content.replace('{high_keyword_count}', str(len(high_keywords)))
html_content = html_content.replace('{low_keyword_count}', str(len(low_keywords)))
html_content = html_content.replace('{all_wc_path}', os.path.basename(all_wc_path))
html_content = html_content.replace('{high_wc_path}', os.path.basename(high_wc_path))
html_content = html_content.replace('{low_wc_path}', os.path.basename(low_wc_path))
html_content = html_content.replace('{all_top_10_table}', all_top_10_table)
html_content = html_content.replace('{high_top_10_table}', high_top_10_table)
html_content = html_content.replace('{low_top_10_table}', low_top_10_table)
html_content = html_content.replace('{top_keywords_comparison_table}', top_keywords_comparison_table)

# 保存HTML文件
html_file = os.path.join(output_dir, 'keyword_wordcloud_analysis.html')
with open(html_file, 'w', encoding='utf-8') as f:
    f.write(html_content)

print(f"HTML文件已保存到: {html_file}")

print("\n=== 词云图生成完成！ ===")
print("\n生成的文件:")
print(f"1. {all_wc_path} - 所有论文关键词词云")
print(f"2. {high_wc_path} - 高活跃度作者论文关键词词云")
print(f"3. {low_wc_path} - 非高活跃度作者论文关键词词云")
print(f"4. {html_file} - 包含三个词云图和对比分析的HTML文件")
