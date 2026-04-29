import openpyxl
import csv
import os
from collections import defaultdict

print("开始最终的作者合作分析...")

# 1. 使用openpyxl直接读取Excel文件，获取作者信息
print("正在读取04表...")
excel_file = r'f:\projects\nlp\NLP-Midterm\r1\csv\04_Diaspora_Articles_12195(1).xlsx'

# 打开Excel文件
# 不使用read_only模式，以便可以使用iter_cols
workbook = openpyxl.load_workbook(excel_file)
sheet = workbook.active

# 获取Diaspora_Author_Names列的索引（根据之前的分析，可能在第76列）
author_col = None
# 只检查前100列，应该足够找到目标列
for row in sheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=100):
    for cell in row:
        if cell.value == 'Diaspora_Author_Names':
            author_col = cell.column
            break
    if author_col:
        break

if author_col is None:
    print("错误：未找到Diaspora_Author_Names列")
    exit(1)

print(f"Diaspora_Author_Names列在第 {author_col} 列")

# 2. 计算每个作者的合作文章数
print("正在计算作者合作文章数...")
author_collab = defaultdict(int)
total_rows = sheet.max_row

for row in range(2, total_rows + 1):  # 从第2行开始，跳过表头
    if row % 1000 == 0:
        print(f"处理了 {row} 行...")
    
    cell_value = sheet.cell(row=row, column=author_col).value
    if cell_value:
        authors = [author.strip() for author in str(cell_value).split(';') if author.strip()]
        if len(authors) > 1:  # 合作文章
            for author in authors:
                author_collab[author] += 1

# 关闭Excel文件
workbook.close()

print(f"已处理 {total_rows - 1} 篇文章")
print(f"已计算 {len(author_collab)} 位作者的合作文章数")

# 3. 读取作者总列表，获取总论文数
print("\n正在读取作者总列表...")
subject_file = r'f:\projects\nlp\NLP-Midterm\r1\csv\unique_authors_with_subject.csv'

# 使用csv模块直接读取
authors_total = {}
with open(subject_file, 'r', encoding='utf-8-sig') as f:
    reader = csv.DictReader(f)
    for row in reader:
        author = row['Author']
        total_papers = int(row['Total_Papers'])
        authors_total[author] = total_papers

print(f"作者总列表包含 {len(authors_total)} 位作者")

# 4. 生成作者合作分析表
print("\n正在生成作者合作分析表...")
output_file = r'f:\projects\nlp\NLP-Midterm\r1\csv\author_collaboration_analysis.csv'

with open(output_file, 'w', newline='', encoding='utf-8-sig') as f:
    writer = csv.writer(f)
    # 写入表头
    writer.writerow(['Author', 'Total_Papers', 'Collaboration_Papers', 'Collaboration_Rate'])
    
    # 写入数据
    for author, total_papers in authors_total.items():
        collab_papers = author_collab.get(author, 0)
        
        if total_papers > 0:
            collab_rate = round(collab_papers / total_papers, 4)
        else:
            collab_rate = 0.0
        
        writer.writerow([author, total_papers, collab_papers, collab_rate])

print(f"作者合作分析表已保存到: {output_file}")

# 5. 区分高活跃度和低活跃度作者
print("\n正在区分高活跃度和低活跃度作者...")

highly_productive = []
low_productive = []

for author, total_papers in authors_total.items():
    collab_papers = author_collab.get(author, 0)
    
    if total_papers > 0:
        collab_rate = round(collab_papers / total_papers, 4)
    else:
        collab_rate = 0.0
    
    if total_papers >= 4:
        highly_productive.append((total_papers, collab_rate))
    else:
        low_productive.append((total_papers, collab_rate))

print(f"高活跃度作者 (≥4篇论文): {len(highly_productive)} 位")
print(f"低活跃度作者 (<4篇论文): {len(low_productive)} 位")

# 6. 计算统计指标
print("\n正在计算统计指标...")

# 高活跃度作者统计
high_rates = [rate for _, rate in highly_productive]
if high_rates:
    high_avg = sum(high_rates) / len(high_rates)
    high_rates_sorted = sorted(high_rates)
    high_med = high_rates_sorted[len(high_rates_sorted) // 2]
    high_max = max(high_rates)
    high_min = min(high_rates)
else:
    high_avg = high_med = high_max = high_min = 0

# 低活跃度作者统计
low_rates = [rate for _, rate in low_productive]
if low_rates:
    low_avg = sum(low_rates) / len(low_rates)
    low_rates_sorted = sorted(low_rates)
    low_med = low_rates_sorted[len(low_rates_sorted) // 2]
    low_max = max(low_rates)
    low_min = min(low_rates)
else:
    low_avg = low_med = low_max = low_min = 0

# 打印比较结果
print("\n=== 高活跃度与低活跃度作者合作率比较 ===")
print(f"{'统计指标':<25} | {'高活跃度作者':<15} | {'低活跃度作者':<15}")
print("-" * 60)
print(f"{'平均合作率':<25} | {high_avg:<15.4f} | {low_avg:<15.4f}")
print(f"{'中位数合作率':<25} | {high_med:<15.4f} | {low_med:<15.4f}")
print(f"{'最高合作率':<25} | {high_max:<15.4f} | {low_max:<15.4f}")
print(f"{'最低合作率':<25} | {high_min:<15.4f} | {low_min:<15.4f}")
print(f"{'作者数量':<25} | {len(highly_productive):<15} | {len(low_productive):<15}")

# 7. 保存对比结果
print("\n正在保存对比结果...")
comparison_file = r'f:\projects\nlp\NLP-Midterm\r1\csv\collaboration_comparison.csv'

with open(comparison_file, 'w', newline='', encoding='utf-8-sig') as f:
    writer = csv.writer(f)
    # 写入表头
    writer.writerow(['Metric', 'Highly Productive Authors', 'Low Productive Authors'])
    
    # 写入数据
    writer.writerow(['Average Collaboration Rate', round(high_avg, 4), round(low_avg, 4)])
    writer.writerow(['Median Collaboration Rate', round(high_med, 4), round(low_med, 4)])
    writer.writerow(['Max Collaboration Rate', round(high_max, 4), round(low_max, 4)])
    writer.writerow(['Min Collaboration Rate', round(high_min, 4), round(low_min, 4)])
    writer.writerow(['Number of Authors', len(highly_productive), len(low_productive)])

print(f"对比结果已保存到: {comparison_file}")

# 8. 显示一些示例数据
print("\n=== 作者合作分析示例 ===")
print("前10位作者的合作情况:")

with open(output_file, 'r', encoding='utf-8-sig') as f:
    reader = csv.reader(f)
    header = next(reader)  # 跳过表头
    print(f"{'Author':<25} {'Total':<8} {'Collab':<8} {'Rate':<8}")
    print("-" * 50)
    
    count = 0
    for row in reader:
        if count < 10:
            print(f"{row[0]:<25} {row[1]:<8} {row[2]:<8} {row[3]:<8}")
            count += 1
        else:
            break

print("\n=== 分析完成！ ===")
print("已生成以下文件:")
print(f"1. {output_file} - 所有作者的合作分析表")
print(f"2. {comparison_file} - 高活跃度与低活跃度作者的对比")
