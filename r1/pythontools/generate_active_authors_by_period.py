import openpyxl
import os
import csv
from collections import defaultdict

print("开始生成active_authors_by_period.csv文件...")

# 1. 从04表中提取作者、年份和论文信息
excel_file = r'f:\projects\nlp\NLP-Midterm\r1\csv\04_Diaspora_Articles_12195(1).xlsx'

# 使用read_only模式打开Excel文件
workbook = openpyxl.load_workbook(excel_file, read_only=True)
sheet = workbook.active

# 找到需要的列
author_col = None
year_col = None
row = sheet[1]  # 获取表头行
for cell in row:
    if cell.value == 'Diaspora_Author_Names':
        author_col = cell.column - 1  # 转换为0-based索引
    elif cell.value == 'Publication Year':
        year_col = cell.column - 1  # 转换为0-based索引

if author_col is None or year_col is None:
    print("错误：未找到所需的列")
    exit(1)

print(f"Diaspora_Author_Names列在索引 {author_col}")
print(f"Publication Year列在索引 {year_col}")

# 收集每个作者在每个年份的论文数
author_year_count = defaultdict(lambda: defaultdict(int))
processed_count = 0
batch_size = 1000
max_rows = sheet.max_row

for row in sheet.iter_rows(min_row=2, max_row=max_rows):
    processed_count += 1
    if processed_count % batch_size == 0:
        print(f"已处理 {processed_count} / {max_rows} 行")
    
    # 获取作者和年份信息
    author_cell = row[author_col].value
    year_cell = row[year_col].value
    
    if author_cell and year_cell:
        # 处理多个作者的情况
        authors = [author.strip() for author in str(author_cell).split(';') if author.strip()]
        year = int(year_cell)
        
        for author in authors:
            author_year_count[author][year] += 1

workbook.close()

print(f"已处理 {processed_count} 行")
print(f"已提取 {len(author_year_count)} 位作者的发表年份数据")

# 2. 生成所有可能的5年期间
# 确定年份范围
all_years = []
for author, year_counts in author_year_count.items():
    all_years.extend(year_counts.keys())

if not all_years:
    print("错误：没有找到任何年份数据")
    exit(1)

min_year = min(all_years)
max_year = max(all_years)

# 生成5年期间列表
periods = []
start_year = 1991
end_year = 2024

current_start = start_year
while current_start + 4 <= end_year:
    period = f"{current_start}-{current_start + 4}"
    periods.append({
        'start': current_start,
        'end': current_start + 4,
        'label': period
    })
    current_start += 1

print(f"生成的5年期间数量: {len(periods)}")
print(f"第一个期间: {periods[0]['label']}")
print(f"最后一个期间: {periods[-1]['label']}")

# 3. 统计每个期间的高活跃度作者数量（在该期间内发表≥5篇）
print("\n统计每个期间的高活跃度作者数量...")

period_stats = []

for period in periods:
    period_start = period['start']
    period_end = period['end']
    period_label = period['label']
    
    active_authors = 0
    
    for author, year_counts in author_year_count.items():
        # 计算该作者在当前期间内的论文总数
        total_in_period = 0
        for year, count in year_counts.items():
            if period_start <= year <= period_end:
                total_in_period += count
        
        # 如果该期间内发表≥5篇，则计数为高活跃度作者
        if total_in_period >= 5:
            active_authors += 1
    
    period_stats.append({
        'period': period_label,
        'count': active_authors
    })
    
    print(f"期间 {period_label}: {active_authors} 位高活跃度作者")

# 4. 保存结果到CSV文件
print("\n保存结果到CSV文件...")
output_file = r'f:\projects\nlp\NLP-Midterm\r1\csv\active_authors_by_period.csv'

with open(output_file, 'w', newline='', encoding='utf-8-sig') as f:
    writer = csv.writer(f)
    # 写入表头
    writer.writerow(['Period', 'Highly Active Authors (>=5 articles)'])
    
    # 写入数据
    for stat in period_stats:
        writer.writerow([stat['period'], stat['count']])

print(f"active_authors_by_period.csv文件已保存到: {output_file}")
print("\n=== 生成完成！ ===")
print(f"共生成了 {len(period_stats)} 个5年期间的统计数据")
