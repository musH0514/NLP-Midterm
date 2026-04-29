import openpyxl
import os
from collections import defaultdict
import numpy as np
import matplotlib.pyplot as plt

print("开始快速分析高活跃度与低活跃度作者的合作差异...")

# 1. 读取高活跃度作者表
print("读取高活跃度作者表...")
high_file = r'f:\projects\nlp\NLP-Midterm\r1\csv\highly_productive_scholars.csv'
high_authors = set()

with open(high_file, 'r', encoding='utf-8-sig') as f:
    next(f)  # 跳过表头
    for line in f:
        parts = line.strip().split(',')
        if parts:
            # 处理带引号的作者名
            if parts[0].startswith('"'):
                # 找到匹配的引号
                author = ','.join(parts[:-3]).strip('"')  # 假设格式是"Author",Articles,Year Range,Average
            else:
                author = parts[0]
            high_authors.add(author)

print(f"高活跃度作者数量: {len(high_authors)}")

# 2. 读取作者学科分类表，获取总论文数和低活跃度作者
print("读取作者学科分类表...")
subject_file = r'f:\projects\nlp\NLP-Midterm\r1\csv\unique_authors_with_subject.csv'

author_total_papers = defaultdict(int)
low_authors = []

with open(subject_file, 'r', encoding='utf-8-sig') as f:
    next(f)  # 跳过表头
    for line in f:
        parts = line.strip().split(',')
        if len(parts) >= 8:
            # 处理带引号的作者名
            if parts[0].startswith('"'):
                # 找到匹配的引号
                author = ','.join(parts[:-7]).strip('"')  # 假设格式是"Author",Main_Subject,...
                total_papers = int(parts[-1])
            else:
                author = parts[0]
                total_papers = int(parts[-1])
            
            author_total_papers[author] = total_papers
            
            # 确定是否为低活跃度作者
            if author not in high_authors and total_papers < 4:
                low_authors.append(author)

print(f"低活跃度作者数量: {len(low_authors)}")
print(f"总作者数: {len(author_total_papers)}")

# 3. 使用openpyxl读取04表，计算合作论文数
print("读取04表，计算合作论文数...")
excel_file = r'f:\projects\nlp\NLP-Midterm\r1\csv\04_Diaspora_Articles_12195(1).xlsx'

# 使用read_only模式打开Excel文件
workbook = openpyxl.load_workbook(excel_file, read_only=True)
sheet = workbook.active

# 找到Diaspora_Author_Names列的索引
author_col = None
row = sheet[1]  # 获取表头行
for cell in row:
    if cell.value == 'Diaspora_Author_Names':
        author_col = cell.column - 1  # 转换为0-based索引
        break

if author_col is None:
    print("错误：未找到Diaspora_Author_Names列")
    exit(1)

print(f"Diaspora_Author_Names列在索引 {author_col}")

# 计算合作论文数
author_collab_papers = defaultdict(int)
processed_count = 0
batch_size = 1000

# 处理所有行
max_rows = sheet.max_row

for row in sheet.iter_rows(min_row=2, max_row=max_rows):
    processed_count += 1
    if processed_count % batch_size == 0:
        print(f"已处理 {processed_count} / {max_rows} 行")
    
    cell_value = row[author_col].value
    if cell_value:
        authors = [author.strip() for author in str(cell_value).split(';') if author.strip()]
        num_authors = len(authors)
        
        if num_authors > 1:  # 合作论文
            for author in authors:
                if author in author_total_papers:  # 只统计我们关注的作者
                    author_collab_papers[author] += 1

workbook.close()

print(f"已处理 {processed_count} 行")
print(f"已计算 {len(author_collab_papers)} 位作者的合作论文数")

# 4. 计算合作率
print("\n计算合作率...")

# 高活跃度作者合作率
high_collab_rates = []
for author in high_authors:
    total = author_total_papers.get(author, 0)
    collab = author_collab_papers.get(author, 0)
    if total > 0:
        high_collab_rates.append(collab / total)

# 低活跃度作者合作率（使用所有低活跃度作者）
low_collab_rates = []
for author in low_authors:
    total = author_total_papers.get(author, 0)
    collab = author_collab_papers.get(author, 0)
    if total > 0:
        low_collab_rates.append(collab / total)

print(f"高活跃度作者有效样本: {len(high_collab_rates)} / {len(high_authors)}")
print(f"低活跃度作者有效样本: {len(low_collab_rates)} / {len(low_authors)}")

# 5. 计算统计指标
print("\n计算统计指标...")

# 高活跃度作者统计
high_stats = {
    'count': len(high_collab_rates),
    'avg_rate': np.mean(high_collab_rates) if high_collab_rates else 0,
    'median_rate': np.median(high_collab_rates) if high_collab_rates else 0
}

# 低活跃度作者统计
low_stats = {
    'count': len(low_collab_rates),
    'avg_rate': np.mean(low_collab_rates) if low_collab_rates else 0,
    'median_rate': np.median(low_collab_rates) if low_collab_rates else 0
}

# 打印统计结果
print("\n=== 高活跃度作者统计 ===")
print(f"作者数量: {high_stats['count']}")
print(f"平均合作率: {high_stats['avg_rate']:.4f}")
print(f"中位数合作率: {high_stats['median_rate']:.4f}")

print("\n=== 低活跃度作者统计 ===")
print(f"作者数量: {low_stats['count']}")
print(f"平均合作率: {low_stats['avg_rate']:.4f}")
print(f"中位数合作率: {low_stats['median_rate']:.4f}")

# 6. 生成对比图表
print("\n生成对比图表...")

# 创建图表目录
output_dir = r'f:\projects\nlp\NLP-Midterm\r1\maps\background'
if not os.path.exists(output_dir):
    os.makedirs(output_dir)

# 图表: 平均合作率对比
plt.figure(figsize=(8, 6))

categories = ['Highly Productive', 'Low Productive']
avg_rates = [high_stats['avg_rate'], low_stats['avg_rate']]

bars = plt.bar(categories, avg_rates, color=['#4ECDC4', '#FF6B6B'], width=0.5)

plt.title('Average Collaboration Rate Comparison', fontsize=16)
plt.ylabel('Collaboration Rate', fontsize=12)
plt.ylim(0, 1)

# 在柱子上显示数值
for bar in bars:
    height = bar.get_height()
    plt.text(bar.get_x() + bar.get_width()/2., height + 0.01,
             f'{height:.4f}', ha='center', va='bottom')

# 添加网格线
plt.grid(axis='y', alpha=0.3)

# 保存图表
chart_file = os.path.join(output_dir, 'collaboration_rate_comparison.png')
plt.savefig(chart_file, dpi=300, bbox_inches='tight')
print(f"图表已保存到: {chart_file}")

# 图表: 合作率分布箱线图
plt.figure(figsize=(10, 6))

# 将数据整理为箱线图需要的格式
data_to_plot = [high_collab_rates, low_collab_rates]

# 创建箱线图
box = plt.boxplot(data_to_plot, patch_artist=True, labels=categories)

# 设置箱子颜色
colors = ['#4ECDC4', '#FF6B6B']
for patch, color in zip(box['boxes'], colors):
    patch.set_facecolor(color)
    patch.set_alpha(0.7)

plt.title('Collaboration Rate Distribution', fontsize=16)
plt.ylabel('Collaboration Rate', fontsize=12)
plt.ylim(0, 1)

# 添加网格线
plt.grid(axis='y', alpha=0.3)

# 保存图表
boxplot_file = os.path.join(output_dir, 'collaboration_rate_distribution.png')
plt.savefig(boxplot_file, dpi=300, bbox_inches='tight')
print(f"箱线图已保存到: {boxplot_file}")

# 7. 生成结果表格
print("\n生成结果表格...")

result_file = r'f:\projects\nlp\NLP-Midterm\r1\csv\collaboration_comparison_result.csv'

with open(result_file, 'w', newline='', encoding='utf-8-sig') as f:
    f.write('Metric,Highly Productive Authors,Low Productive Authors\n')
    f.write(f'Number of Authors,{high_stats["count"]},{low_stats["count"]}\n')
    f.write(f'Average Collaboration Rate,{high_stats["avg_rate"]:.4f},{low_stats["avg_rate"]:.4f}\n')
    f.write(f'Median Collaboration Rate,{high_stats["median_rate"]:.4f},{low_stats["median_rate"]:.4f}\n')

print(f"结果表格已保存到: {result_file}")

print("\n分析完成！")
print("\n生成的文件:")
print(f"1. {chart_file} - 平均合作率对比图")
print(f"2. {boxplot_file} - 合作率分布箱线图")
print(f"3. {result_file} - 对比结果表格")
print("\n注意：本次分析使用了完整的xlsx表格数据和所有低活跃度作者")
print("分析结果基于全部可用数据")
